"""
Generate SAP Free Courses Excel from SAP Learning Hub data.
Covers: Finance (FI), Controlling (CO), Material Management (MM), Sales & Distribution (SD),
        Supply Chain Management (SCM), SAP BTP, CPI/Integration Suite, SAP Build
Columns: id, url (Enable Now only), role (sub-topic area), title, module, description (short), lastUpdated, sapHelpLink
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os

# ── Course data gathered from SAP Learning Hub (free courses only) ──────────
# role = functional sub-area / topic (e.g., Asset, GL Accounting, Payables)
# module = SAP module (Finance (FI), Controlling (CO), Material Management (MM), etc.)
# description = short ~10-15 words on what you do/learn

courses = [
    # ═══════════════════════════════════════════════════════════════════
    # MODULE: Finance (FI)
    # ═══════════════════════════════════════════════════════════════════
    {
        "id": "FI-001",
        "url": "",
        "role": "Financial Overview",
        "title": "Discovering Finance in SAP S/4HANA",
        "module": "Finance (FI)",
        "description": "Overview of Finance capabilities in SAP S/4HANA, 8 courses, 14 hr+",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/learning-journeys/discovering-finance-in-sap-s-4hana"
    },
    {
        "id": "FI-002",
        "url": "",
        "role": "Financial Accounting",
        "title": "Implementing Financial Accounting in SAP S/4HANA Cloud",
        "module": "Finance (FI)",
        "description": "Implement Financial Accounting in S/4HANA Cloud Public Edition, leads to certification",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/learning-journeys/implementing-financial-accounting-in-sap-s4hana-cloud"
    },
    {
        "id": "FI-003",
        "url": "",
        "role": "Management Accounting",
        "title": "Implementing Management Accounting in SAP S/4HANA Cloud",
        "module": "Finance (FI)",
        "description": "Implement Management Accounting in S/4HANA Cloud, leads to certification",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/learning-journeys/implementing-management-accounting-in-sap-s-4hana-cloud"
    },
    {
        "id": "FI-004",
        "url": "",
        "role": "Financial Accounting",
        "title": "Outlining the Financial Accounting Overview in SAP S/4HANA",
        "module": "Finance (FI)",
        "description": "Learn different parts of Financial Accounting with demos and exercises",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/learning-journeys/outlining-the-financial-accounting-overview-in-sap-s-4hana"
    },
    {
        "id": "FI-005",
        "url": "",
        "role": "GL Accounting",
        "title": "Outlining the Record to Report Process in SAP S/4HANA",
        "module": "Finance (FI)",
        "description": "Key concepts of Financial Accounting master data and postings in S/4HANA",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/learning-journeys/outlining-the-record-to-report-process-in-sap-s-4hana"
    },
    {
        "id": "FI-006",
        "url": "",
        "role": "GL Accounting",
        "title": "Designing the Record to Report Process in SAP S/4HANA",
        "module": "Finance (FI)",
        "description": "Configure Accruals Management, Financial Statement Reporting and Financial Closing",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/learning-journeys/designing-the-record-to-report-process-in-sap-s-4hana"
    },
    {
        "id": "FI-007",
        "url": "",
        "role": "Payables",
        "title": "Describing the Payables Management Process in SAP S/4HANA",
        "module": "Finance (FI)",
        "description": "Fundamental processes in Payables Management on SAP S/4HANA",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/learning-journeys/describing-the-payables-management-process-in-sap-s-4hana"
    },
    {
        "id": "FI-008",
        "url": "",
        "role": "Payables",
        "title": "Using the Payables Management Process in SAP S/4HANA",
        "module": "Finance (FI)",
        "description": "Payment processing, configuring and executing the payment program",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/learning-journeys/using-the-payables-management-process-in-sap-s-4hana"
    },
    {
        "id": "FI-009",
        "url": "",
        "role": "Receivables",
        "title": "Describing the Receivables Management Process in SAP S/4HANA",
        "module": "Finance (FI)",
        "description": "Fundamental business processes in Receivables Management area on S/4HANA",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/learning-journeys/describing-the-receivables-management-process-in-sap-s-4hana"
    },
    {
        "id": "FI-010",
        "url": "",
        "role": "Asset",
        "title": "Designing the Asset Accounting Process in SAP S/4HANA",
        "module": "Finance (FI)",
        "description": "Configure asset accounting, run integrated processes, year-end closing",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/learning-journeys/designing-the-asset-accounting-process"
    },
    {
        "id": "FI-011",
        "url": "",
        "role": "Financial Close",
        "title": "Exploring Accounting & Financial Close",
        "module": "Finance (FI)",
        "description": "Explore the accounting and financial close process in SAP",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/learning-journeys/exploring-accounting-financial-close"
    },
    {
        "id": "FI-012",
        "url": "",
        "role": "Real Estate",
        "title": "Describing Contracts for Real Estate Management in SAP S/4HANA",
        "module": "Finance (FI)",
        "description": "Manage Real Estate contracts in SAP S/4HANA across deployment options",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/learning-journeys/describing-contracts-for-real-estate-management-in-sap-s4hana"
    },
    {
        "id": "FI-013",
        "url": "",
        "role": "Profitability",
        "title": "Detailing Profitability Accounting for Discrete Industries",
        "module": "Finance (FI)",
        "description": "Configure Profitability Accounting in Make-to-Stock scenarios, hands-on (F2231)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/detailing-profitability-accounting-for-discrete-industries"
    },
    {
        "id": "FI-014",
        "url": "",
        "role": "Profitability",
        "title": "Detailing Profitability Accounting in Make-to-Order Scenarios",
        "module": "Finance (FI)",
        "description": "Configure Profitability Accounting in Make-to-Order scenarios, hands-on (F2241)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/detailing-profitability-accounting-in-make-to-order-scenarios"
    },
    {
        "id": "FI-015",
        "url": "",
        "role": "Consolidation",
        "title": "Rule-based Consolidation of Investments",
        "module": "Finance (FI)",
        "description": "Implement rule-based consolidations for S/4HANA Cloud group reporting (F9631)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/rule-based-consolidation-of-investments"
    },
    {
        "id": "FI-016",
        "url": "",
        "role": "Profit Center",
        "title": "Detailing Profit Center Reorganization",
        "module": "Finance (FI)",
        "description": "Analyze and perform organizational changes using reorganization tools (F1261)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/detailing-profit-center-reorganization"
    },
    {
        "id": "FI-017",
        "url": "",
        "role": "Treasury",
        "title": "Discovering Treasury Management in SAP S/4HANA",
        "module": "Finance (FI)",
        "description": "Capabilities and business benefits for efficient Treasury operations (F4011)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/discovering-treasury-management"
    },
    {
        "id": "FI-018",
        "url": "",
        "role": "Bank Accounting",
        "title": "Administration of Bank Accounts",
        "module": "Finance (FI)",
        "description": "Bank account administration core business processes in S/4HANA (F4111)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/administration-of-bank-accounts"
    },
    {
        "id": "FI-019",
        "url": "",
        "role": "Consolidation",
        "title": "Outlining SAP S/4HANA Cloud for Group Reporting",
        "module": "Finance (FI)",
        "description": "Discover financial consolidation business processes in SAP (F9111)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/outlining-sap-s-4hana-cloud-for-group-reporting"
    },
    {
        "id": "FI-020",
        "url": "",
        "role": "Consolidation",
        "title": "Implementing SAP S/4HANA Cloud for Group Reporting",
        "module": "Finance (FI)",
        "description": "Configure S/4HANA Cloud for group reporting and financial consolidation",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/learning-journeys/implementing-sap-s-4hana-cloud-for-group-reporting"
    },
    {
        "id": "FI-021",
        "url": "",
        "role": "Consolidation",
        "title": "Performing Consolidation with SAP S/4HANA Cloud for Group Reporting",
        "module": "Finance (FI)",
        "description": "Perform consolidation and analyze consolidated values in group reporting",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/learning-journeys/performing-consolidation-with-sap-s-4hana-cloud-for-group-reporting"
    },
    {
        "id": "FI-022",
        "url": "",
        "role": "Working Capital",
        "title": "Exploring SAP Taulia Working Capital Management",
        "module": "Finance (FI)",
        "description": "Scope and importance of SAP Taulia Working Capital Management (SL_FIN511C)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/exploring-sap-taulia-working-capital-management-1"
    },
    {
        "id": "FI-023",
        "url": "",
        "role": "AI in Finance",
        "title": "Discovering SAP Business AI Capabilities for SAP Finance",
        "module": "Finance (FI)",
        "description": "SAP Business AI strategy for financial management and data-driven decisions",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/learning-journeys/discovering-new-ai-capabilities-for-sap-finance"
    },
    {
        "id": "FI-024",
        "url": "",
        "role": "Billing",
        "title": "Discovering SAP Solutions for Quote-to-Cash Management",
        "module": "Finance (FI)",
        "description": "End-to-end Solution and Subscription-based business process (BR473)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/discovering-sap-solutions-for-quote-to-cash-management-public-cloud-for-business-user"
    },
    {
        "id": "FI-025",
        "url": "",
        "role": "Localization",
        "title": "Exploring Localization as a Self-Service for SAP S/4HANA Cloud",
        "module": "Finance (FI)",
        "description": "Capabilities of Localization as a Self-Service for implementing localizations (S4CLS)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/applying-localization-as-a-self-service-for-sap-s-4hana-cloud-public-edition-1"
    },

    # ═══════════════════════════════════════════════════════════════════
    # MODULE: Controlling (CO)
    # ═══════════════════════════════════════════════════════════════════
    {
        "id": "CO-001",
        "url": "",
        "role": "Cost Management",
        "title": "Outlining Cost Management and Profitability Analysis",
        "module": "Controlling (CO)",
        "description": "Core business processes in Overhead Cost Accounting (F2111)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/outlining-cost-management-and-profitability-analysis"
    },
    {
        "id": "CO-002",
        "url": "",
        "role": "Overhead Cost",
        "title": "Performing Overhead Cost Controlling in SAP S/4HANA",
        "module": "Controlling (CO)",
        "description": "Cost centers, projects, and internal orders in Overhead Cost Accounting",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/learning-journeys/performing-overhead-cost-controlling-in-sap-s-4hana"
    },
    {
        "id": "CO-003",
        "url": "",
        "role": "Overhead Cost",
        "title": "Evaluating Overhead Cost Accounting in SAP S/4HANA",
        "module": "Controlling (CO)",
        "description": "Concepts, terminology, and configuration of Overhead Cost Accounting",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/learning-journeys/evaluating-overhead-cost-accounting-in-sap-s-4hana"
    },
    {
        "id": "CO-004",
        "url": "",
        "role": "Production Cost",
        "title": "Evaluating Production Accounting in Make-to-Stock Scenarios",
        "module": "Controlling (CO)",
        "description": "Production Accounting by Order and Period using cost objects like production orders",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/learning-journeys/evaluating-production-accounting-in-make-to-stock-scenarios-in-sap-s-4hana"
    },
    {
        "id": "CO-005",
        "url": "",
        "role": "Overhead Cost",
        "title": "Detailing Posting Control, Allocations, and Settlement",
        "module": "Controlling (CO)",
        "description": "Implement Overhead Cost Accounting posting and settlement in S/4HANA (F2211)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/detailing-posting-control-allocations-and-settlement"
    },

    # ═══════════════════════════════════════════════════════════════════
    # MODULE: Material Management (MM)
    # ═══════════════════════════════════════════════════════════════════
    {
        "id": "MM-001",
        "url": "",
        "role": "Procurement",
        "title": "Implementing Sourcing and Procurement (Public Cloud)",
        "module": "Material Management (MM)",
        "description": "Implement sourcing and procurement processes in S/4HANA Cloud Public Edition",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/learning-journeys/implement-sap-s-4hana-cloud-public-edition-for-sourcing-and-procurement"
    },
    {
        "id": "MM-002",
        "url": "",
        "role": "Procurement",
        "title": "Implementing Sourcing and Procurement (Private Cloud)",
        "module": "Material Management (MM)",
        "description": "Implement sourcing and procurement processes in S/4HANA Private Cloud",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/learning-journeys/implementing-sap-s-4hana-cloud-private-edition-sourcing-and-procurement"
    },
    {
        "id": "MM-003",
        "url": "",
        "role": "Implementation",
        "title": "Implementing SAP S/4HANA Cloud Public Edition",
        "module": "Material Management (MM)",
        "description": "Fundamental knowledge to implement S/4HANA Cloud Public Edition (S4C01)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/implementing-sap-s-4hana-cloud-public-edition"
    },
    {
        "id": "MM-004",
        "url": "",
        "role": "End-to-End Process",
        "title": "Exploring End-to-End Business Processes in SAP Business Suite",
        "module": "Material Management (MM)",
        "description": "Integrated end-to-end business processes in SAP S/4HANA Cloud (IEE2E)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/exploring-end-to-end-business-processes-in-sap-business-suite"
    },

    # ═══════════════════════════════════════════════════════════════════
    # MODULE: Sales and Distribution (SD)
    # ═══════════════════════════════════════════════════════════════════
    {
        "id": "SD-001",
        "url": "",
        "role": "Sales",
        "title": "SAP S/4HANA Sales Insights",
        "module": "Sales and Distribution (SD)",
        "description": "Introduction to fundamental processes in SAP S/4HANA Sales (S4690)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/sap-s-4hana-sales-insights"
    },
    {
        "id": "SD-002",
        "url": "",
        "role": "Sales",
        "title": "Exploring SAP S/4HANA Sales Essentials",
        "module": "Sales and Distribution (SD)",
        "description": "Execute essential process steps in SAP S/4HANA Sales, hands-on (S46000)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/exploring-sap-s-4hana-sales-essentials"
    },
    {
        "id": "SD-003",
        "url": "",
        "role": "Sales",
        "title": "Implementing Sales in SAP S/4HANA Cloud Public Edition",
        "module": "Sales and Distribution (SD)",
        "description": "Preparation for S/4HANA Cloud Sales certification, 7 courses, 49 hr+",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/learning-journeys/implementing-sap-s4hana-cloud-public-edition-sales"
    },
    {
        "id": "SD-004",
        "url": "",
        "role": "Service Cloud",
        "title": "Configuring SAP Sales and Service Cloud as Administrator",
        "module": "Sales and Distribution (SD)",
        "description": "System setup, fine tuning, personalizing features and Go Live preparation",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/learning-journeys/sap-sales-and-service-cloud-administration"
    },

    # ═══════════════════════════════════════════════════════════════════
    # MODULE: Supply Chain Management (SCM)
    # ═══════════════════════════════════════════════════════════════════
    {
        "id": "SCM-001",
        "url": "",
        "role": "Supply Chain",
        "title": "Positioning SAP Supply Chain Management Solutions",
        "module": "Supply Chain Management (SCM)",
        "description": "Master SAP SCM solutions within SAP Business Suite, end-to-end processes",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/learning-journeys/positioning-sap-supply-chain-solutions"
    },
    {
        "id": "SCM-002",
        "url": "",
        "role": "Supply Chain",
        "title": "Discovering SAP Supply Chain Management Solutions",
        "module": "Supply Chain Management (SCM)",
        "description": "End-to-end SCM solutions addressing COO key priorities (SL_SCM100)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/discovering-sap-supply-chain-management-solutions"
    },
    {
        "id": "SCM-003",
        "url": "",
        "role": "Supply Chain",
        "title": "Positioning SAP SCM Solutions (Course)",
        "module": "Supply Chain Management (SCM)",
        "description": "Position SAP SCM solutions within SAP Business Suite context (SL_SCM200)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/positioning-sap-supply-chain-management-solutions"
    },
    {
        "id": "SCM-004",
        "url": "",
        "role": "AI in SCM",
        "title": "Introducing SAP Business AI for Supply Chain Management",
        "module": "Supply Chain Management (SCM)",
        "description": "Strategic AI applications and practical use cases in supply chain (SL_SCM150)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/introducing-sap-business-ai-for-sap-supply-chain-management"
    },
    {
        "id": "SCM-005",
        "url": "",
        "role": "Warehouse",
        "title": "Discovering Extended Warehouse Management with SAP S/4HANA",
        "module": "Supply Chain Management (SCM)",
        "description": "Warehouse management deployment options and EWM functions in S/4HANA",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/learning-journeys/discovering-extended-warehouse-management-with-sap-s-4hana"
    },
    {
        "id": "SCM-006",
        "url": "",
        "role": "Business Network",
        "title": "Configuring SAP Business Network Supply Chain Collaboration",
        "module": "Supply Chain Management (SCM)",
        "description": "Configuration and administration of Business Network SC Collaboration Add-On",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/learning-journeys/configuring-sap-business-network-supply-chain-collaboration-add-on"
    },

    # ═══════════════════════════════════════════════════════════════════
    # MODULE: SAP BTP (Business Technology Platform)
    # ═══════════════════════════════════════════════════════════════════
    {
        "id": "BTP-001",
        "url": "",
        "role": "Platform",
        "title": "Exploring SAP Business Technology Platform",
        "module": "SAP BTP",
        "description": "SAP BTP strategy, app development, integration, data, analytics and AI (BTP100)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/exploring-sap-business-technology-platform"
    },
    {
        "id": "BTP-002",
        "url": "",
        "role": "Platform",
        "title": "Discovering SAP Business Technology Platform",
        "module": "SAP BTP",
        "description": "Discover SAP Business Technology Platform capabilities (SL_BTP400)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/discovering-sap-business-technology-platform-1"
    },
    {
        "id": "BTP-003",
        "url": "",
        "role": "Architecture",
        "title": "Becoming an SAP BTP Solution Architect",
        "module": "SAP BTP",
        "description": "Design intelligent enterprise solutions using SAP BTP, leads to certification",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/learning-journeys/becoming-an-sap-btp-solution-architect"
    },
    {
        "id": "BTP-004",
        "url": "",
        "role": "Administration",
        "title": "Operating SAP Business Technology Platform",
        "module": "SAP BTP",
        "description": "Set up and configure SAP BTP platform administration (BTP200)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/operating-sap-business-technology-platform"
    },
    {
        "id": "BTP-005",
        "url": "",
        "role": "Cloud Development",
        "title": "Developing Applications for SAP BTP Cloud Foundry",
        "module": "SAP BTP",
        "description": "Develop and run apps on SAP BTP Cloud Foundry runtime (CF400)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/developing-applications-for-sap-btp-cloud-foundry-runtime"
    },
    {
        "id": "BTP-006",
        "url": "",
        "role": "CAP Development",
        "title": "Develop Extensions with CAP (BTP Developer's Guide)",
        "module": "SAP BTP",
        "description": "Build a CAP application following the SAP BTP Developer's Guide (CLD200)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/develop-extensions-with-cap-following-the-sap-btp-developer-s-guide"
    },
    {
        "id": "BTP-007",
        "url": "",
        "role": "Security",
        "title": "Architecting Security for SAP BTP",
        "module": "SAP BTP",
        "description": "IAM, role-based access, API security, network security in hybrid SAP (SECCL3)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/architecting-security-for-sap-business-technology-platform"
    },
    {
        "id": "BTP-008",
        "url": "",
        "role": "Security",
        "title": "Introducing SAP Cloud Identity Services",
        "module": "SAP BTP",
        "description": "Identity Authentication Service and Identity Provisioning Service (SECCL1)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/introducing-sap-cloud-identity-services"
    },
    {
        "id": "BTP-009",
        "url": "",
        "role": "Operations",
        "title": "Operating with SAP Cloud ALM",
        "module": "SAP BTP",
        "description": "SAP Cloud ALM monitoring, analysis, Business Service and Event Management (CALM40)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/operating-with-sap-cloud-alm"
    },
    {
        "id": "BTP-010",
        "url": "",
        "role": "Cloud Development",
        "title": "Developing Advanced Extensions with SAP Cloud SDK",
        "module": "SAP BTP",
        "description": "Side-by-side extensions on SAP BTP for S/4HANA Cloud with Cloud SDK",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/learning-journeys/develop-advanced-extensions-with-sap-cloud-sdk"
    },
    {
        "id": "BTP-011",
        "url": "",
        "role": "ABAP",
        "title": "Learning the Basics of ABAP Programming on SAP BTP",
        "module": "SAP BTP",
        "description": "Basic ABAP programming for on-premise and cloud system environments",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/learning-journeys/learn-the-basics-of-abap-programming-on-sap-btp"
    },
    {
        "id": "BTP-012",
        "url": "",
        "role": "ABAP",
        "title": "Introducing SAP ABAP Platform Fundamentals",
        "module": "SAP BTP",
        "description": "ABAP Platform concepts and tools for system administration",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/learning-journeys/introducing-sap-abap-platform-fundamentals"
    },
    {
        "id": "BTP-013",
        "url": "",
        "role": "ABAP",
        "title": "Setting up an ABAP Environment on SAP BTP",
        "module": "SAP BTP",
        "description": "Step-by-step guidance to set up ABAP Environment in SAP BTP Account",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/learning-journeys/setting-up-an-abap-environment-on-sap-btp"
    },
    {
        "id": "BTP-014",
        "url": "",
        "role": "AI",
        "title": "Introduction to AI Core",
        "module": "SAP BTP",
        "description": "Use SAP AI Core to execute and operate artificial intelligence assets (AICO01)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/introduction-to-ai-core"
    },
    {
        "id": "BTP-015",
        "url": "",
        "role": "AI",
        "title": "Discovering SAP Business AI",
        "module": "SAP BTP",
        "description": "SAP Business AI for automating processes and enhancing decision-making (SL_BAI100)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/discovering-sap-business-ai"
    },

    # ═══════════════════════════════════════════════════════════════════
    # MODULE: CPI / Integration Suite
    # ═══════════════════════════════════════════════════════════════════
    {
        "id": "CPI-001",
        "url": "",
        "role": "Integration",
        "title": "Developing with SAP Integration Suite",
        "module": "CPI / Integration Suite",
        "description": "Create APIs, iFlows, Cloud Integration and API Management development",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/learning-journeys/developing-with-sap-integration-suite"
    },
    {
        "id": "CPI-002",
        "url": "",
        "role": "Integration",
        "title": "Accelerate Enterprise Integrations with SAP Integration Suite",
        "module": "CPI / Integration Suite",
        "description": "End-to-end integration, expose APIs and integrate heterogeneous systems (MOOC_cp10)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/accelerate-enterprise-integrations-with-sap-integration-suite"
    },
    {
        "id": "CPI-003",
        "url": "",
        "role": "Migration",
        "title": "SAP Process Orchestration to Integration Suite Migration",
        "module": "CPI / Integration Suite",
        "description": "Migrate from SAP PO / PI to SAP Integration Suite (PO2IS)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/sap-process-orchestration-to-sap-integration-suite-migration"
    },
    {
        "id": "CPI-004",
        "url": "",
        "role": "Administration",
        "title": "Administering SAP Integration Suite",
        "module": "CPI / Integration Suite",
        "description": "Essential admin tasks, setup, and management of Integration Suite (CLD920)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/administering-sap-integration-suite"
    },
    {
        "id": "CPI-005",
        "url": "",
        "role": "Connectivity",
        "title": "Connecting SAP BTP with On-Premise via Cloud Connector",
        "module": "CPI / Integration Suite",
        "description": "Connect SAP BTP applications with on-premise systems using Cloud Connector (ADM002)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/connecting-sap-btp-with-on-premise-via-cloud-connector"
    },
    {
        "id": "CPI-006",
        "url": "",
        "role": "Automation",
        "title": "Discovering Enterprise Automation with SAP",
        "module": "CPI / Integration Suite",
        "description": "SAP Signavio, SAP Build, and Integration Suite for process optimization (SEA100)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/discovering-enterprise-automation-with-sap"
    },
    {
        "id": "CPI-007",
        "url": "",
        "role": "Process Orchestration",
        "title": "Discovering SAP Process Orchestration",
        "module": "CPI / Integration Suite",
        "description": "Introduction to SAP PO and Advanced Adapter Engine Extended (BIT100)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/discovering-sap-process-orchestration"
    },
    {
        "id": "CPI-008",
        "url": "",
        "role": "Event Mesh",
        "title": "Discovering Event-Driven Integration with Advanced Event Mesh",
        "module": "CPI / Integration Suite",
        "description": "Event streaming and event management for event-driven architecture (CLD905)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/discovering-event-driven-integration-with-sap-integration-suite-advanved-event-mesh"
    },
    {
        "id": "CPI-009",
        "url": "",
        "role": "Web Services",
        "title": "Developing SOAP Web Services on SAP ERP",
        "module": "CPI / Integration Suite",
        "description": "Build SOAP web services on SAP ERP and web service framework (BIT102)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/devoloping-soap-web-services-on-sap-erp"
    },
    {
        "id": "CPI-010",
        "url": "",
        "role": "Advisory",
        "title": "Getting Started with SAP Integration Solution Advisory",
        "module": "CPI / Integration Suite",
        "description": "Define and execute enterprise integration strategy using ISAM (CLD910)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/getting-started-with-sap-integration-solution-advisory-methodology"
    },
    {
        "id": "CPI-011",
        "url": "",
        "role": "Hybrid Integration",
        "title": "Accelerating Hybrid Integrations on RedHat OpenShift",
        "module": "CPI / Integration Suite",
        "description": "Edge Integration Cell: design in the cloud, deploy on-premise (HIRED1)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/accelerating-hybrid-integrations-with-sap-integration-suite-on-redhat-openshift"
    },
    {
        "id": "CPI-012",
        "url": "",
        "role": "IDoc/RFC",
        "title": "Developing Integration Scenarios using IDoc/RFC Adapter",
        "module": "CPI / Integration Suite",
        "description": "Create iFlows using IDoc and BAPI adapter in Process Orchestration (BIT101)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/developing-integration-scenarios-using-idoc-rfc-adapter-of-sap-process-orchestration"
    },
    {
        "id": "CPI-013",
        "url": "",
        "role": "Process Orchestration",
        "title": "Developing Business Processes with SAP Process Orchestration",
        "module": "CPI / Integration Suite",
        "description": "Build business processes and integration scenarios with SAP PO",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/learning-journeys/developing-business-processes-with-sap-process-orchestration"
    },
    {
        "id": "CPI-014",
        "url": "",
        "role": "Commerce Integration",
        "title": "Integrate SAP S/4HANA Cloud with SAP Commerce Cloud",
        "module": "CPI / Integration Suite",
        "description": "Implement integration of SAP Commerce Cloud with S/4HANA Cloud backend",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/learning-journeys/implement-an-integration-of-sap-s-4hana-cloud-with-sap-commerce-cloud"
    },

    # ═══════════════════════════════════════════════════════════════════
    # MODULE: SAP Build
    # ═══════════════════════════════════════════════════════════════════
    {
        "id": "BUILD-001",
        "url": "",
        "role": "Low-Code",
        "title": "Experiencing End-To-End SAP Build",
        "module": "SAP Build",
        "description": "Build Apps, Process Automation, and Work Zone end-to-end experience (BTP04)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/experiencing-end-to-end-sap-build"
    },
    {
        "id": "BUILD-002",
        "url": "",
        "role": "App Development",
        "title": "Developing with SAP Build - From Apps to Automation",
        "module": "SAP Build",
        "description": "Develop apps and processes with SAP Build, leads to certification",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/learning-journeys/developing-with-sap-build-from-apps-to-automation"
    },
    {
        "id": "BUILD-003",
        "url": "",
        "role": "Low-Code",
        "title": "Develop and Automate with SAP Build",
        "module": "SAP Build",
        "description": "Low-code: create apps, automate processes, design business sites (BTP110)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/develop-and-automate-with-sap-build"
    },
    {
        "id": "BUILD-004",
        "url": "",
        "role": "Process Automation",
        "title": "Developing with SAP Build Process Automation",
        "module": "SAP Build",
        "description": "Build automations and processes, reuse content packages (SPA400)",
        "lastUpdated": "01/15/2025",
        "sapHelpLink": "https://learning.sap.com/courses/developing-with-sap-build-process-automation"
    },
]


def create_excel(data, output_path):
    """Create a professionally formatted Excel workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SAP Free Courses"

    # ── Header styling ──────────────────────────────────────────────
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["id", "url", "role", "title", "module", "description", "lastUpdated", "sapHelpLink"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # ── Data rows ───────────────────────────────────────────────────
    data_font = Font(name="Calibri", size=10)
    data_align = Alignment(vertical="top", wrap_text=True)
    link_font = Font(name="Calibri", size=10, color="0563C1", underline="single")

    alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

    for row_idx, course in enumerate(data, 2):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()
        for col_idx, key in enumerate(headers, 1):
            value = course.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border
            cell.fill = fill

            # Make sapHelpLink clickable
            if key == "sapHelpLink" and value:
                cell.hyperlink = value
                cell.font = link_font
            # Make url (Enable Now) clickable if present
            if key == "url" and value:
                cell.hyperlink = value
                cell.font = link_font

    # ── Column widths ───────────────────────────────────────────────
    col_widths = {
        "A": 12,   # id
        "B": 20,   # url (Enable Now)
        "C": 22,   # role (sub-topic area)
        "D": 55,   # title
        "E": 30,   # module
        "F": 65,   # description (shorter now)
        "G": 14,   # lastUpdated
        "H": 70,   # sapHelpLink
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # ── Freeze header row ───────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Auto-filter ─────────────────────────────────────────────────
    ws.auto_filter.ref = f"A1:H{len(data) + 1}"

    # ── Save ────────────────────────────────────────────────────────
    wb.save(output_path)
    print(f"Excel saved to: {output_path}")
    print(f"Total courses: {len(data)}")

    # Module summary
    modules = {}
    for c in data:
        m = c["module"]
        modules[m] = modules.get(m, 0) + 1
    print("\nCourses per module:")
    for m, count in modules.items():
        print(f"  {m}: {count}")

    # Role/Sub-topic summary
    roles = {}
    for c in data:
        r = c["role"]
        roles[r] = roles.get(r, 0) + 1
    print("\nCourses per role (sub-topic):")
    for r, count in sorted(roles.items()):
        print(f"  {r}: {count}")


if __name__ == "__main__":
    # Try OneDrive Desktop first, fallback to regular Desktop
    desktop = os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop")
    if not os.path.exists(desktop):
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    os.makedirs(desktop, exist_ok=True)
    output = os.path.join(desktop, "SAP_Free_Courses_Catalog.xlsx")
    create_excel(courses, output)
